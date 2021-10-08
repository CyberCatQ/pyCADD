import os
import sys
import openpyxl
import re
import pandas as pd
from openpyxl.styles import Font

root_path = os.path.abspath(os.path.dirname(__file__)).split('src')[0]  # 项目路径 绝对路径
result_path = root_path + 'lib/result/'                                 # 对接结果数据文件库
info_path = root_path + 'lib/info/'                                     # 晶体基本信息数据文件库
pdb_path = root_path.split('automatedMD')[0]                            # PDB项目绝对路径(如果有)
inputfile_path = root_path + 'lib/list/'                                # py4schrodinger输入文件库

font = Font(name='等线',size=12)

gene_dict = {'NR1A1': 'TR_alpha',
             'NR1A2': 'TR_beta',
             'NR1B1': 'RAR_alpha',
             'NR1B2': 'RAR_beta',
             'NR1B3': 'RAR_gamma',
             'NR1C1': 'PPAR_alpha',
             'NR1C2': 'PPAR_beta',
             'NR1C3': 'PPAR_gamma',
             'NR1D1': 'Rev-erb',
             'NR1F1': 'ROR_alpha',
             'NR1F2': 'ROR_beta',
             'NR1F3': 'ROR_gamma',
             'NR1H3': 'LXR_alpha',
             'NR1H2': 'LXR_beta',
             'NR1H4': 'FXR_alpha',
             'NR1H5': 'FXR_beta',
             'NR1I1': 'VDR',
             'NR1I2': 'PXR',
             'NR1I3': 'CAR',
             'NR2A1': 'HNF4_alpha',
             'NR2A2': 'HNF4_gamma',
             'NR2B1': 'RXR_alpha',
             'NR2B2': 'RXR_beta',
             'NR2B3': 'RXR_gamma',
             'NR2C1': 'TR2',
             'NR2C2': 'TR4',
             'NR2E2': 'TLL',
             'NR2E3': 'PNR',
             'NR2F1': 'COUP-TFI',
             'NR2F2': 'COUP-TFII',
             'NR2F6': 'EAR2',
             'NR3A1': 'ER_alpha',
             'NR3A2': 'ER_beta',
             'NR3B1': 'ERR_alpha',
             'NR3B2': 'ERR_beta',
             'NR3B3': 'ERR_gamma',
             'NR3C1': 'GR',
             'NR3C2': 'MR',
             'NR3C3': 'PR',
             'NR3C4': 'AR',
             'NR4A1': 'NGFIB',
             'NR4A2': 'NURR1',
             'NR4A3': 'NOR1',
             'NR5A1': 'SF1',
             'NR5A2': 'LRH1',
             'NR6A1': 'GCNF',
             'NR0B1': 'DAX1',
             'NR0B2': 'SHP'}

def result_merge(ligname=None):
    '''
    合并对接结果数据并输出到同一EXCEL中 并产生一张汇总sheet

    Parameter
    ----------
    ligname : str
        外源配体名称(如果有)
    '''
    precision = input('Enter the precision of docking:(SP|XP)').strip().upper()
    withlig = '_' + precision
    if ligname:
        withlig = '_' + ligname + '_' + precision
    mergefile = pdb_path + 'DOCK_FINAL_RESULTS%s.xlsx' % withlig
    data_list = []  # 汇总数据表

    resultfiles = os.popen("cd %s && ls | grep 'RESULTS%s.csv'" % (result_path, withlig)).read().splitlines()
    writer = pd.ExcelWriter(mergefile)
    for resultfile in resultfiles:
        gene = resultfile.split('_')[0]
        resultdata = pd.read_csv(result_path + resultfile)
        resultdata.to_excel(writer, gene, index=False)
        
        resultdata['GENE'] = gene
        data_list.append(resultdata)
    
    data_total = pd.concat(data_list)
    data_total.to_excel(writer, 'TOTAL', index=False)

    writer.save()

    wb = openpyxl.load_workbook(mergefile)
    for ws in wb:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
    wb.save(mergefile)

def info_merge():
    '''
    合并晶体基础信息并分不同sheet输出到同一EXCEL中 并产生一张汇总sheet
    '''
    mergefile = pdb_path + 'CRYSTALS_INFO.xlsx'
    infofiles = os.popen('cd %s && ls' % info_path).read().splitlines()
    writer = pd.ExcelWriter(mergefile)
    data_list = []  # 汇总数据表

    for infofile in infofiles:
        gene = infofile.split('.')[0]
        data = pd.read_csv(info_path + infofile)
        data.to_excel(writer, gene, index=False)

        data['GENE'] = gene
        data_list.append(data)

    data_total = pd.concat(data_list)
    data_total.to_excel(writer, 'TOTAL', index=False)
    writer.save()

    wb = openpyxl.load_workbook(mergefile)
    for ws in wb:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
    wb.save(mergefile)

def _identify_gen(df, key_col1=None, key_col2=None):
    '''
    生成唯一识别号IDENTIFY列确保数据合并正确
    将直接修改传入的Dataframe

    Parameters
    ----------
    df : pandas.Dataframe
        要处理的Dataframe数据
    key_col1 : str
        用于生成识别号的第一个列名
    key_col2 : str
        用于生成识别号的第二个列名
    
    '''
    
    for index, row in df.iterrows():
        df.at[index, 'IDENTIFY'] = str(row[key_col1]) + str(row[key_col2])

def merge_info_result(ligname:str=None):
    '''
    将晶体基础信息与对接结果合并为总EXCEL

    ligname : str
        外源配体名称（如果有）
    '''


    precision = input('Enter the precision of docking:(SP|XP)').strip().upper()
    withlig = '_' + precision
    if ligname:
        withlig = '_' + ligname + '_' + precision

    mergefile = pdb_path + 'FINAL_RESULTS%s.xlsx' % withlig      # 总EXCEL文件名
    writer = pd.ExcelWriter(mergefile)

    info_file_name = input('Enter the PATH of Info Database File:') # 手动输入晶体信息文件名称(可能随版本迭代改变)
    if not os.path.isfile(info_file_name):
        raise FileNotFoundError('No such file or directory: %s' % info_file_name)
    infofile = os.path.abspath(info_file_name)

    withlig = '_' + precision
    if ligname:
        withlig = '_' + ligname + '_' + precision
    resultfile = pdb_path + 'DOCK_FINAL_RESULTS%s.xlsx' % withlig    

    try:                                                    # 读取晶体信息
        infos = pd.read_excel(infofile, 'TOTAL')
    except ValueError:
        infos = pd.read_excel(infofile)

    results = pd.read_excel(resultfile, 'TOTAL')            # 读取对接结果

    # 生成用于匹配的唯一识别号
    _identify_gen(infos, 'PDB ID', 'Gene Name')
    _identify_gen(results, 'PDB', 'GENE')

    tmp_merge_data = pd.merge(left=infos, right=results, how='left', on='IDENTIFY')
    merge_data = tmp_merge_data[['Name', 'Abbreviation', 'Gene Name', 'PDB ID', 'Ligand_x', 'Ligand ID','Ligand_y',
       'Docking_Score', 'rmsd', 'precision', 'MMGBSA_dG_Bind', 'Volume', 'Site_Score','Comformation', 'Title', 'Reference', 'DOI', 'Times Cited']]
    merge_data = merge_data.rename(columns={'Ligand_x':'Origin Ligand', 'Ligand ID':'Origin Ligand ID', 'Ligand_y':'Docking Ligand'})
    merge_data.to_excel(writer, index=False)
    writer.save()

    wb = openpyxl.load_workbook(mergefile)
    for ws in wb:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
    wb.save(mergefile)

def pivottable_gen(result_file:str=None, key1:str='Comformation', key2:str='Gene Name'):
    '''
    生成数据透视表

    Parameter
    ----------
    result_file : str
        计算结果文件
    key1 : str
        透视表行名称 默认构象类型
    key2 : str
        透视表列名称 默认基因名称
        
    '''
    if not result_file:
        result_file = input('Enter the PATH of FINAL_RESULTS File:').strip()
        
    prefix = os.path.basename(result_file).split('.')[0]
    property_list = ['Docking_Score', 'Volume', 'MMGBSA_dG_Bind', 'rmsd']       # 要考察的核心数据
    data = pd.read_excel(result_file)
    pivottable_file = pdb_path + '/' + prefix + '_Pivot_table.xlsx'

    writer = pd.ExcelWriter(pivottable_file)

    for property in property_list:
        pvt = data.pivot_table(property, index=key1, columns=key2, aggfunc='mean')
        pvt.to_excel(writer, sheet_name= property + '_mean')
    
    writer.save()

    wb = openpyxl.load_workbook(pivottable_file)
    for ws in wb:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
    wb.save(pivottable_file)

def cal_ave_min(lib_path:str=None, dock_path:str=None, property:str='Docking_Score'):
    '''
    计算绝对均值、相对均值
    Parameters
    ----------
    lib_path : str
        原始数据库文件路径
    dock_path : str
        外源配体对接结果文件路径
    property : str
        要计算的属性名称 默认对接分数Docking Score
    '''
    def merge_df(df1, df2, suffixes=('_x','_y')):
        merge = pd.merge(left=df1, right=df2, on='Gene Name', how='left', suffixes=suffixes)
        return merge
    
    if not lib_path:
        lib_path = input('Enter the path of Docking Library File:').strip()
    if not dock_path:
        dock_path = input('Enter the path of EX-Ligand Result File:').strip()
    
    prefix = os.path.basename(dock_path).split('.')[0]
    ave_min_file = pdb_path + '/' + prefix + '_Pivot_table_' + property + '.xlsx'

    path = os.path.abspath(lib_path)
    raw_data = pd.read_excel(path)
    agonist_data = raw_data[raw_data['Comformation'] == 'Agonist']
    antagonist_data = raw_data[raw_data['Comformation'] == 'Antagonist']

    lig_file = os.path.basename(dock_path)
    lig_name = lig_file.split('_')[2]

    dock_data = pd.read_excel(dock_path)
    dock_data = dock_data[dock_data['Docking Ligand'] == lig_name]
    # 去重必须小心在不同基因成员中的同一PDB ID被删除
    dock_data.drop_duplicates(subset = ['Gene Name', 'Origin Ligand ID', 'PDB ID', 'Docking_Score'], keep='first', inplace=True)

    ave_all = raw_data.pivot_table(property, index= 'Gene Name')
    ave_agonist = agonist_data.pivot_table(property, index='Gene Name')
    ave_antagonist = antagonist_data.pivot_table(property, index='Gene Name')
    ave_merge = merge_df(merge_df(ave_all, ave_agonist, ('_all','_agonist')),ave_antagonist)
    ave_merge.rename(columns={property:'%s_antagonist' % property}, inplace=True)

    min_all = raw_data[['Gene Name', property]].groupby(by='Gene Name').min()
    min_agonist = agonist_data[['Gene Name', property]].groupby(by='Gene Name').min()
    min_antagonist = antagonist_data[['Gene Name', property]].groupby(by='Gene Name').min()
    min_merge = merge_df(merge_df(min_all, min_agonist, ('_all', '_agonist')),min_antagonist)
    min_merge.rename(columns={property :'%s_antagonist' % property}, inplace=True)

    _hit = {}
    total_gene_count = raw_data[raw_data['Docking_Score'].notnull()]['Gene Name'].value_counts().to_dict()
    dock_gene_count = dock_data['Gene Name'].value_counts().to_dict()
    for index, value in dock_gene_count.items():
        _hit[index] = value / total_gene_count[index]
    hit_percent = pd.DataFrame.from_dict(_hit, orient='index', columns=['Hit'])
    hit_percent.reset_index().rename(columns={'index':'Gene Name'})
    hit_percent.sort_values(by='Hit', ascending=False, inplace=True)

    writer = pd.ExcelWriter(ave_min_file)
    for data in [ave_merge, min_merge, hit_percent]:
        for index, row in data.iterrows():
            data.loc[index, 'Abbreviation'] = gene_dict[index]

    ave_merge.to_excel(writer, sheet_name='AVERAGE')
    min_merge.to_excel(writer, sheet_name='Min')
    hit_percent.to_excel(writer, sheet_name='Hit Percent')

    for index, row in dock_data.iterrows():
        dock_data.loc[index, 'AVERAGE_All'] = row[property] / ave_all.loc['%s' % row['Gene Name']][property]
        dock_data.loc[index, 'AVERAGE_Agonist'] = row[property] / ave_agonist.loc['%s' % row['Gene Name']][property]
        try:
            dock_data.loc[index, 'AVERAGE_Antagonist'] = row[property] / ave_antagonist.loc['%s' % row['Gene Name']][property]
        except KeyError:
            continue
    
    for index, row in dock_data.iterrows():
        dock_data.loc[index, 'Min_All'] = row[property] / min_all.loc['%s' % row['Gene Name']][property]
        dock_data.loc[index, 'Min_Agonist'] = row[property] / min_agonist.loc['%s' % row['Gene Name']][property]
        try:
            dock_data.loc[index, 'Min_Antagonist'] = row[property] / min_antagonist.loc['%s' % row['Gene Name']][property]
        except KeyError:
            continue

    _identify_gen(dock_data, 'Gene Name', 'PDB ID')
    _identify_gen(raw_data, 'Gene Name', 'PDB ID')

    for index, row in dock_data.iterrows():
        dock_data.loc[index, 'Abbreviation'] = gene_dict[row['Gene Name']]
        dock_data.loc[index, 'PDB_%s' % property] = row[property] / float(raw_data[raw_data['IDENTIFY'] == row['IDENTIFY']][property])
    
    dock_data.drop('IDENTIFY',axis=1, inplace=True)
    dock_data.to_excel(writer, sheet_name=lig_name, index=False)
    writer.save()

    wb = openpyxl.load_workbook(ave_min_file)
    for ws in wb:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
    wb.save(ave_min_file)

def process_listfile():
    '''
    解析已存在的py4schrodinger输入文件 确保配体与目标基因的蛋白关联

    Return
    ---------
    dic
        { GENE: ['PDBID1,ligname1,agonist', 'PDBID2,ligname2,', ...] }
    '''
    list_dic = {}
    inputfiles = os.popen('cd %s && ls' % inputfile_path).read().splitlines()
    for inputfile in inputfiles:
        gene = inputfile.split('.')[0]
        with open(inputfile_path + inputfile) as f:
            cps = f.read().splitlines()
        list_dic[gene] = cps

    return list_dic

def conform_classify():
    '''
    晶体构象分类(Agonist/Antagonist)
    '''
    conform_dic = process_listfile()

    agonist_match = re.compile('agonist', re.IGNORECASE)
    antagonist_match = re.compile('antagonist', re.IGNORECASE)
    agonist_list = []
    antagonist_list = []
    agonist_ref = []
    antagonist_ref = []

    classifyfile = pdb_path + 'COMFORM_CLASSIFY.xlsx'
    infofiles = os.popen('cd %s && ls' % info_path).read().splitlines()
    writer = pd.ExcelWriter(classifyfile)

    for infofile in infofiles:
        gene = infofile.split('.')[0]
        data = pd.read_csv(info_path + infofile)
        agonist_ID = []
        antagonist_ID = []

        for cp in conform_dic[gene]:                # 根据文件中的flag识别为激动剂或拮抗剂
            _flag = cp.split(',')[2]
            _pdb = cp.split(',')[0]
            _lig = cp.split(',')[1]

            if re.search(antagonist_match, _flag):
                antagonist_ID.append(_pdb)
            elif re.search(agonist_match, _flag):
                agonist_ID.append(_pdb)

        # 解析CSV
        for index, row in data.iterrows():
            if row['PDBID'] in antagonist_ID:      # 拮抗剂命中
                antagonist_list.append({'Gene Name' : gene, 'PDB ID' : row['PDBID'],'Conformation' : 'Antagonist','Title' : row['title'], 'Reference' : row['reference'], 'DOI' : row['DOI']})
                antagonist_ref.append(row['DOI'])
            elif row['PDBID'] in agonist_ID:          # 激动剂命中
                agonist_list.append({'Gene Name' : gene, 'PDB ID' : row['PDBID'],'Conformation' : 'Agonist','Title' : row['title'], 'Reference' : row['reference'], 'DOI' : row['DOI']})
                agonist_ref.append(row['DOI'])

    agonist_df = pd.DataFrame(agonist_list)
    antagonist_df = pd.DataFrame(antagonist_list)
    agonist_df.to_excel(writer, 'Agonist', index=False)
    antagonist_df.to_excel(writer, 'Antagonist', index=False)

    # 文献去重
    antagonist_ref = set(antagonist_ref)
    agonist_ref = set(agonist_ref)
    antagonist_ref_df = pd.DataFrame(antagonist_ref)
    agonist_ref_df = pd.DataFrame(agonist_ref)
    agonist_ref_df.to_excel(writer, 'Agonist_Reference', index=False)
    agonist_ref_df.to_csv(pdb_path + 'Agonist_Ref.csv', index=False)
    antagonist_ref_df.to_excel(writer, 'Antagonist_Reference', index=False)
    antagonist_ref_df.to_csv(pdb_path + 'Antagonist_Ref.csv', index=False)
    writer.save()
    
    wb = openpyxl.load_workbook(classifyfile)
    for ws in wb:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
    wb.save(classifyfile)

def main():
    print('''
    Script for Merging Data after Docking.

1. Merge the basic information of all crystals in automatedMD/lib/info
2. Merge all docking results data in automatedMD/lib/result
3. Merge all docking with exogenous ligand results data in automatedMD/lib/result
4. Merge docking results to crystals info
5. Merge docking with exogenous ligand results to crystals info
6. Classfify all crystals via title or reference
7. Pivot Table File Generation
8. Calculate Average and Minimum Value of Result

0. Exit
''')
    print('Enter the code for the operation to be performed:')
    _flag = input()
    if _flag == '1':
        info_merge()
    elif _flag == '2':
        result_merge()
    elif _flag == '3':
        ligname = input('Enter the ligand name:')
        result_merge(ligname)
    elif _flag == '4':
        merge_info_result()
    elif _flag == '5':
        ligname = input('Enter the ligand name:')
        merge_info_result(ligname=ligname)
    elif _flag == '6':
        conform_classify()
    elif _flag == '7':
        pivottable_gen()
    elif _flag == '8':
        lib_path = os.path.abspath(pdb_path + 'FINAL_RESULTS_SP.xlsx')
        if not os.path.exists(lib_path):
            lib_path = input('Enter the path of Docking Library File:').strip()
        print('Using Essential Data:', lib_path)
        dock_path = input('Enter the path of EX-Ligand Result File:').strip()
        cal_ave_min(lib_path, dock_path, property='Docking_Score')
        cal_ave_min(lib_path, dock_path, property='MMGBSA_dG_Bind')
    else:
        sys.exit()
    
if __name__ == '__main__':
    main()